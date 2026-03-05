"""
Excel Reporter - Generación de reportes en Excel
"""
import pandas as pd
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import PieChart, Reference
import os

class ExcelReporter:
    """Agente para generar reportes Excel del monitoreo"""
    
    def __init__(self, output_dir='reports'):
        self.output_dir = output_dir
        os.makedirs(output_dir, exist_ok=True)
    
    def generate_hourly_report(self, articles, stats, hour_label=None):
        """Genera reporte de la hora"""
        if hour_label is None:
            hour_label = datetime.now().strftime('%Y%m%d_%H%M')
        
        filename = f"{self.output_dir}/reporte_{hour_label}.xlsx"
        
        wb = Workbook()
        
        # Hoja 1: Resumen
        ws_summary = wb.active
        ws_summary.title = "Resumen"
        self._create_summary_sheet(ws_summary, stats)
        
        # Hoja 2: Noticias
        ws_news = wb.create_sheet("Noticias")
        self._create_news_sheet(ws_news, articles)
        
        # Hoja 3: Alertas
        ws_alerts = wb.create_sheet("Alertas")
        alerts = [a for a in articles if a.get('is_alert')]
        self._create_alerts_sheet(ws_alerts, alerts)
        
        # Hoja 4: Análisis de Sentimiento
        ws_sentiment = wb.create_sheet("Sentimiento")
        self._create_sentiment_sheet(ws_sentiment, articles)
        
        wb.save(filename)
        print(f"[Excel] Reporte guardado: {filename}")
        return filename
    
    def _create_summary_sheet(self, ws, stats):
        """Crea hoja de resumen"""
        # Título
        ws['A1'] = "REPORTE DE MONITOREO - Paloma Valencia"
        ws['A1'].font = Font(size=16, bold=True, color="FFFFFF")
        ws['A1'].fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        ws.merge_cells('A1:D1')
        ws['A1'].alignment = Alignment(horizontal='center')
        
        # Fecha
        ws['A3'] = "Generado:"
        ws['B3'] = datetime.now().strftime('%d/%m/%Y %H:%M')
        ws['A3'].font = Font(bold=True)
        
        # Estadísticas
        row = 5
        headers = ['Métrica', 'Valor', 'Porcentaje']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col)
            cell.value = header
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.alignment = Alignment(horizontal='center')
        
        data = [
            ('Total Noticias', stats.get('total', 0), '100%'),
            ('Positivas', stats.get('positive', 0), f"{stats.get('positive', 0)/max(stats.get('total', 1), 1)*100:.1f}%"),
            ('Negativas', stats.get('negative', 0), f"{stats.get('negative', 0)/max(stats.get('total', 1), 1)*100:.1f}%"),
            ('Neutrales', stats.get('neutral', 0), f"{stats.get('neutral', 0)/max(stats.get('total', 1), 1)*100:.1f}%"),
            ('Alertas', stats.get('alerts', 0), f"{stats.get('alerts', 0)/max(stats.get('total', 1), 1)*100:.1f}%"),
        ]
        
        for i, (metric, value, pct) in enumerate(data, 1):
            ws.cell(row=row+i, column=1, value=metric)
            ws.cell(row=row+i, column=2, value=value)
            ws.cell(row=row+i, column=3, value=pct)
            
            # Colores según tipo
            if 'Positivas' in metric:
                ws.cell(row=row+i, column=2).fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            elif 'Negativas' in metric:
                ws.cell(row=row+i, column=2).fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            elif 'Alertas' in metric:
                ws.cell(row=row+i, column=2).fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        
        # Ajustar anchos
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15
    
    def _create_news_sheet(self, ws, articles):
        """Crea hoja de noticias"""
        headers = ['Fecha', 'Hora', 'Título', 'Fuente', 'Sentimiento', 'Relevancia', 'Alerta', 'URL']
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        
        for i, article in enumerate(articles, 2):
            date_str = article.get('published_at', '')
            if date_str:
                try:
                    dt = datetime.fromisoformat(date_str.replace('Z', '+00:00'))
                    ws.cell(row=i, column=1, value=dt.strftime('%d/%m/%Y'))
                    ws.cell(row=i, column=2, value=dt.strftime('%H:%M'))
                except:
                    ws.cell(row=i, column=1, value='N/A')
                    ws.cell(row=i, column=2, value='N/A')
            
            ws.cell(row=i, column=3, value=article.get('title', ''))
            ws.cell(row=i, column=4, value=article.get('source', ''))
            
            sentiment = article.get('sentiment', 'neutral')
            cell_sent = ws.cell(row=i, column=5, value=sentiment.upper())
            if sentiment == 'positive':
                cell_sent.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            elif sentiment == 'negative':
                cell_sent.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            
            ws.cell(row=i, column=6, value=article.get('relevance_score', 0))
            ws.cell(row=i, column=7, value='SÍ' if article.get('is_alert') else 'NO')
            ws.cell(row=i, column=8, value=article.get('url', ''))
        
        # Ajustar anchos
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 8
        ws.column_dimensions['C'].width = 50
        ws.column_dimensions['D'].width = 20
        ws.column_dimensions['E'].width = 12
        ws.column_dimensions['F'].width = 12
        ws.column_dimensions['G'].width = 8
        ws.column_dimensions['H'].width = 40
    
    def _create_alerts_sheet(self, ws, alerts):
        """Crea hoja de alertas"""
        ws['A1'] = "ALERTAS DETECTADAS"
        ws['A1'].font = Font(size=14, bold=True, color="FFFFFF")
        ws['A1'].fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
        ws.merge_cells('A1:F1')
        
        headers = ['Fecha/Hora', 'Título', 'Fuente', 'Razón', 'Relevancia', 'URL']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        
        for i, alert in enumerate(alerts, 4):
            ws.cell(row=i, column=1, value=alert.get('published_at', '')[:16])
            ws.cell(row=i, column=2, value=alert.get('title', ''))
            ws.cell(row=i, column=3, value=alert.get('source', ''))
            ws.cell(row=i, column=4, value=alert.get('alert_reason', ''))
            ws.cell(row=i, column=5, value=alert.get('relevance_score', 0))
            ws.cell(row=i, column=6, value=alert.get('url', ''))
        
        for col in ['A', 'B', 'C', 'D', 'E', 'F']:
            ws.column_dimensions[col].width = 20 if col != 'B' else 50
    
    def _create_sentiment_sheet(self, ws, articles):
        """Crea análisis de sentimiento"""
        ws['A1'] = "ANÁLISIS DE SENTIMIENTO"
        ws['A1'].font = Font(size=14, bold=True)
        
        # Contar por sentimiento
        positive = [a for a in articles if a.get('sentiment') == 'positive']
        negative = [a for a in articles if a.get('sentiment') == 'negative']
        neutral = [a for a in articles if a.get('sentiment') == 'neutral']
        
        ws['A3'] = "Sentimiento"
        ws['B3'] = "Cantidad"
        ws['C3'] = "Porcentaje"
        
        for cell in [ws['A3'], ws['B3'], ws['C3']]:
            cell.font = Font(bold=True)
        
        data = [
            ('Positivo', len(positive)),
            ('Negativo', len(negative)),
            ('Neutral', len(neutral))
        ]
        
        total = sum(d[1] for d in data)
        for i, (sent, count) in enumerate(data, 4):
            ws.cell(row=i, column=1, value=sent)
            ws.cell(row=i, column=2, value=count)
            ws.cell(row=i, column=3, value=f"{count/max(total,1)*100:.1f}%")
        
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 12
    
    def generate_daily_summary(self, db_connection):
        """Genera resumen diario consolidado"""
        today = datetime.now().strftime('%Y%m%d')
        filename = f"{self.output_dir}/resumen_diario_{today}.xlsx"
        
        # TODO: Consultar DB y generar resumen
        print(f"[Excel] Resumen diario: {filename}")
        return filename
